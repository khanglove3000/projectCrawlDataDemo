from selenium import webdriver
from time import sleep
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook

def get_value_excel(filename, cellname):
	wb = load_workbook(filename)
	Sheet1 = wb['Sheet1']
	wb.close()
	return Sheet1[cellname].value()

def update_value_excel(filename, cellname, value):
	wb = openpyxl.load_workbook(filename)
	Sheet1 = wb['Sheet1']
	Sheet1[cellname].value = value
	wb.close()
	wb.save(filename)

col_name_mst = "A"
col_name_cty = "B"
col_name_boss = "C"
col_name_phone = "D"
filename = "file.xlsx"

url = "https://masothue.com/"

for i_row in range(2, 10):
	cell_name_mst = "%s%s"%(col_name_mst,i_row)
	cell_name_cty = "%s%s"%(col_name_cty,i_row)
	cell_name_boss = "%s%s"%(col_name_boss,i_row)
	cell_name_phone = "%s%s"%(col_name_phone,i_row)
	print(cell_name_mst)
	mst = get_value_excel(filename, cell_name_mst)
	driver = webdriver.Chrome("C:\\Users\\khang\\Desktop\\crawl_data\chromedriver")
	driver.get(url)



